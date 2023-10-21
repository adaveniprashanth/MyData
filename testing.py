try:
    import configparser
    import pandas as pd
    import numpy as np
    import xlsxwriter
    import xlrd
    from openpyxl import load_workbook,Workbook
    from openpyxl.styles import PatternFill,Alignment
    from openpyxl.utils import get_column_letter
    import sys,re
    from datetime import date
    import time
    from selenium import webdriver
    from selenium.webdriver.common.by import By
except ModuleNotFoundError:
    print("you have to install the below packages to run")
    print("pandas,numpy,openpyxl and xlrd")
    print("pip install pandas\npip install numpy\npip install openpyxl\npip install xlrd")
    print("other modules are xlsxwriter,openpyxl and selenium")
    print("python script and input file should be in same folder")
    sys.exit("install all modules")


#importing from the config gile
config=configparser.ConfigParser()
try:
    config.read("configurations.ini") #reading the config file
except FileNotFoundError:
    sys.exit("file not present. please check file")

input_file= config["FILESettings"]["input_filename"]
approved_consolidated_filename=config["FILESettings"]["approved_consolidated_filename"]
not_approved_consolidated_jiras=config["FILESettings"]["not_approved_consolidated_jiras"]
jira_submission_file = config["FILESettings"]["jira_submission_filename"]
workweek=config["FILESettings"]["work_week"]
PONumber=config["FILESettings"]["PONumber"]
Vendor=config["FILESettings"]["Vendor"]
Team = config["FILESettings"]["Team"]
Platform = config["FILESettings"]["Platform"]
SKU = config["FILESettings"]["SKU"]
BillableHeader =config["FILESettings"]["BillableHeader"]
Location = config["FILESettings"]["Location"]
L1Approver =config["FILESettings"]["L1Approver"]
L2Approver = config["FILESettings"]["L2Approver"]
column_width= config["FILESettings"]["column_width"]
c1_billing=int(config["FILESettings"]["c1_billing"])
c2_billing=int(config["FILESettings"]["c2_billing"])
c3_billing=int(config["FILESettings"]["c3_billing"])
c4_billing=int(config["FILESettings"]["c4_billing"])
c5_billing=int(config["FILESettings"]["c5_billing"])
jira_update=int(config["FILESettings"]["Jira_update"])

def update_story_points(jira_ID,expected_point):
    driver = webdriver.Chrome("C:\chromedriver\chromedriver.exe")
    url="https://jira.devtools.intel.com/browse/"
    expected_story_points = expected_point  # find from excel and convert to int
    jira_id = jira_ID
    driver.get("{}{}".format(url,jira_id))
    driver.maximize_window()
    time.sleep(10)
    story_points = int(driver.find_element(By.XPATH, '//div[@class="value type-float editable-field inactive"]').text)
    if story_points != expected_story_points:
        print("update story points")
        driver.find_element(By.XPATH, '//span [@class="icon aui-icon aui-icon-small aui-iconfont-edit"]').click()
        time.sleep(10)
        driver.find_element(By.ID, "customfield_11204").clear()
        driver.find_element(By.ID, "customfield_11204").send_keys(expected_story_points)
        driver.find_element(By.ID, "edit-issue-submit").click()
    else:
        print("story already points set to {}".format(expected_story_points))

#convert column numbers to column names
def colToExcel(col): # col is 1 based
    excelCol = str()
    div = col
    while div:
        (div, mod) = divmod(div-1, 26) # will return (x, 0 .. 25)
        excelCol = chr(mod + 65) + excelCol
    return excelCol

def beautify_excel(file_name,rows_length,column_length):    
    wb = load_workbook(file_name) 
    ws = wb["data_set"]
    
    #ws.column_dimensions['A'].width = 25
    
    #applying the colour to the column headings and adjusting the columns width 
    fill_cell = PatternFill(patternType='solid', fgColor='ffff00')
    for i in range(1,column_length+1):
        ws.cell(row=1,column=i).fill =fill_cell
        ws.column_dimensions[colToExcel(i)].width = column_width
    
    #adding hyperlink
    for i in range(2,rows_length+2):
        ws.cell(row=i, column=6).hyperlink = "https://jira.devtools.intel.com/browse/"+ws.cell(row=i, column=6).value
        ws.cell(row=i, column=6).value = ws.cell(row=i, column=6).value
        ws.cell(row=i, column=6).style = "Hyperlink"
    
    #Align the cells to center
    for i in range(1,rows_length+2):
        for j in range(1,column_length+1):
            ws.cell(row=i,column=j).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

    wb.save(file_name)

file = open("Log.txt",'w')

if 1:
    input_filename=input_file#'WW_40_43_1.xlsx'
    consolidate_approved_jiras=approved_consolidated_filename#'consollidated_bill.xlsx'
    try:
        xL = pd.ExcelFile(input_filename)        
    except PermissionError:
        sys.exit("close the opened Excel Sheet"+input_filename)
    list_of_sheets =xL.sheet_names    
    print("total no.of sheets are ",len(list_of_sheets)) 
    print("sheets in excel file are\n",xL.sheet_names)
    file.write("Total sheets in {} excel file are {}\n".format(input_filename,len(list_of_sheets)))
    file.write("sheets in {} excel file are\n{}\n".format(input_filename,xL.sheet_names))
    
    
    df = pd.read_excel(input_filename,sheet_name=xL.sheet_names)#accessing sheets by name
    
    jira_approved=pd.DataFrame()
    jira_not_approved=pd.DataFrame()
    complexity_dict = {}
    changed_jiras=0
    
    for i in list_of_sheets:
        print("sheet name is ",i)
        l=len(df[i].iloc[:])
        #print("total rows in {0} shett are {1}".format(i,len(df[i].iloc[:,:])))
        for j in range(l):
            print("assignee is ",df[i].iloc[j,0])
            #print("jira is ",df[i].iloc[j,9])
            if pd.notna(df[i].loc[j,"assignee".lower()]) and pd.notna(df[i].iloc[j,"Intel Leads Approval".lower()]) and df[i].iloc[j,"Intel Leads Approval".lower()].strip().lower() == "approved":
                #if pd.notna(df[i].loc[j,"Remarks"]) and jira_update:
                #    update_story_points(df[i].loc[j,"Key"],int(re.findall("[0-9]",df[i].loc[j,"Remarks"])[0]))
                #    changed_jiras+=1
                print("approved jira is ",df[i].iloc[j,9])
