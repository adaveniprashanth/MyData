try:
    import configparser
    import pandas as pd
    import numpy as np
    import xlsxwriter
    import xlrd
    from openpyxl import load_workbook,Workbook #For reference -->https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.fonts.html
    from openpyxl.styles import PatternFill,Alignment,Font
    from openpyxl.utils import get_column_letter
    import sys,re
    from datetime import date
    import time
    from selenium import webdriver
    from selenium.webdriver.common.by import By
except ModuleNotFoundError:
    print("you have to install the below packages to run")
    print("pandas,numpy,openpyxl and xlrd")
    print("pip install pandas\npip install numpy\npip install openpyxl\npip install xlrd")
    print("other modules are xlsxwriter,openpyxl and selenium")
    print("python script and input file should be in same folder")
    print("column names spelling should be like this:Assignee	Issue Type	Project	Reporter	Sprint	Status	Story Points	Summary	Domain	Key	Labels	Complexity	Intel Leads Approval	Remarks")
    sys.exit("install all modules")


#importing from the config gile
config=configparser.ConfigParser()
try:
    config.read("configurations.ini") #reading the config file
except FileNotFoundError:
    sys.exit("file not present. please check file")

input_file= config["FILESettings"]["input_filename"]
approved_consolidated_filename=config["FILESettings"]["approved_consolidated_filename"]
not_approved_consolidated_jiras=config["FILESettings"]["not_approved_consolidated_jiras"]
jira_submission_file = config["FILESettings"]["jira_submission_filename"]
workweek=config["FILESettings"]["work_week"]
PONumber=config["FILESettings"]["PONumber"]
Vendor=config["FILESettings"]["Vendor"]
Team = config["FILESettings"]["Team"]
Platform = config["FILESettings"]["Platform"]
SKU = config["FILESettings"]["SKU"]
BillableHeader =config["FILESettings"]["BillableHeader"]
Location = config["FILESettings"]["Location"]
L1Approver =config["FILESettings"]["L1Approver"]
L2Approver = config["FILESettings"]["L2Approver"]
column_width= config["FILESettings"]["column_width"]
c1_billing=int(config["FILESettings"]["c1_billing"])
c2_billing=int(config["FILESettings"]["c2_billing"])
c3_billing=int(config["FILESettings"]["c3_billing"])
c4_billing=int(config["FILESettings"]["c4_billing"])
c5_billing=int(config["FILESettings"]["c5_billing"])
jira_update=int(config["FILESettings"]["Jira_update"])

def update_story_points(jira_ID,expected_point):
    driver = webdriver.Chrome("C:\chromedriver\chromedriver.exe")
    url="https://jira.devtools.intel.com/browse/"
    expected_story_points = expected_point  # find from excel and convert to int
    jira_id = jira_ID
    driver.get("{}{}".format(url,jira_id))
    driver.maximize_window()
    time.sleep(10)
    story_points = int(driver.find_element(By.XPATH, '//div[@class="value type-float editable-field inactive"]').text)
    if story_points != expected_story_points:
        print("update story points")
        driver.find_element(By.XPATH, '//span [@class="icon aui-icon aui-icon-small aui-iconfont-edit"]').click()
        time.sleep(10)
        driver.find_element(By.ID, "customfield_11204").clear()
        driver.find_element(By.ID, "customfield_11204").send_keys(expected_story_points)
        driver.find_element(By.ID, "edit-issue-submit").click()
    else:
        print("story already points set to {}".format(expected_story_points))

#convert column numbers to column names
def colToExcel(col): # col is 1 based
    excelCol = str()
    div = col
    while div:
        (div, mod) = divmod(div-1, 26) # will return (x, 0 .. 25)
        excelCol = chr(mod + 65) + excelCol
    return excelCol

def beautify_excel(file_name,rows_length,column_length):    
    wb = load_workbook(file_name) 
    ws = wb["data_set"]
    
    #ws.column_dimensions['A'].width = 25
    
    #applying the colour to the column headings and adjusting the columns width 
    fill_cell = PatternFill(patternType='solid', fgColor='ffff00')
    for i in range(1,column_length+1):
        ws.cell(row=1,column=i).fill =fill_cell
        ws.column_dimensions[colToExcel(i)].width = column_width
    
    #adding hyperlink
    for i in range(2,rows_length+2):
        ws.cell(row=i, column=6).hyperlink = "https://jira.devtools.intel.com/browse/"+ws.cell(row=i, column=6).value
        ws.cell(row=i, column=6).value = ws.cell(row=i, column=6).value
        ws.cell(row=i, column=6).style = "Hyperlink"
    
    #Align the cells to center
    for i in range(1,rows_length+2):
        for j in range(1,column_length+1):
            ws.cell(row=i,column=j).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

    wb.save(file_name)

file = open("Log.txt",'w')

if 1:
    input_filename=input_file#'WW_40_43_1.xlsx'
    consolidate_approved_jiras=approved_consolidated_filename#'consollidated_bill.xlsx'
    try:
        xL = pd.ExcelFile(input_filename)        
    except PermissionError:
        sys.exit("close the opened Excel Sheet"+input_filename)
    list_of_sheets =xL.sheet_names    
    print("total no.of sheets are ",len(list_of_sheets)) 
    print("sheets in excel file are\n",xL.sheet_names)
    file.write("Total sheets in {} excel file are {}\n".format(input_filename,len(list_of_sheets)))
    file.write("sheets in {} excel file are\n{}\n".format(input_filename,xL.sheet_names))
    
    
    df = pd.read_excel(input_filename,sheet_name=xL.sheet_names)#accessing sheets by name
    
    jira_approved=pd.DataFrame()
    jira_not_approved=pd.DataFrame()
    complexity_dict = dict.fromkeys(('C1','C2','C3','C4','C5'),0)
    changed_jiras=0
    
    for i in list_of_sheets:
        print("sheet name is ",i)
        print("columns names in {} sheet are {} ".format(i, df[i].columns.values))
        l=len(df[i].iloc[:])
        print("total rows in {0} shett are {1}".format(i,len(df[i].iloc[:,:])))
        for j in range(l):
            if pd.notna(df[i].loc[j,"Assignee"]) and pd.notna(df[i].loc[j,"Intel Leads Approval"]) and df[i].loc[j,"Intel Leads Approval"].strip().lower() == "approved":
                if pd.notna(df[i].loc[j,"Remarks"]) and jira_update:
                   update_story_points(df[i].loc[j,"Key"],int(re.findall("[0-9]",df[i].loc[j,"Remarks"])[0]))
                   changed_jiras+=1
                print("approved jira is ",df[i].loc[j,"Key"])
                file.write("approved jira in sheet {} is {}\n".format(i,df[i].loc[j,"Key"]))
                d = { "Assignee":df[i].loc[j,"Assignee"],
                      "Issue Type":df[i].loc[j,"Issue Type"],
                      "Story Points":df[i].loc[j,"Story Points"],
                      "Summary":df[i].loc[j,"Summary"],
                      "Domain":i,
                      "Key":df[i].loc[j,"Key"],
                      "Complexity":df[i].loc[j,"Complexity"]
                    }
                ser=pd.Series(d)
                jira_approved=jira_approved.append(ser,ignore_index=True)
                #print("dict is",complexity_dict)
                complexity_dict[df[i].loc[j,"Complexity"]]=complexity_dict.get(df[i].loc[j,"Complexity"],0)+1
                
            elif pd.notna(df[i].loc[j,"Assignee"]):
                print("not approved jira is ",df[i].loc[j,"Key"])
                d = { "Assignee":df[i].loc[j,"Assignee"],
                      "Issue Type":df[i].loc[j,"Issue Type"],
                      "Story Points":df[i].loc[j,"Story Points"],
                      "Summary":df[i].loc[j,"Summary"],
                      "Domain":i,
                      "Key":df[i].loc[j,"Key"],
                      "Complexity":df[i].loc[j,"Complexity"]
                    }
                file.write("not approved jira in sheet {} is {}\n".format(i,df[i].loc[j,"Key"]))
                ser=pd.Series(d)
                jira_not_approved=jira_not_approved.append(ser,ignore_index=True)
        
    print("approved jira is",jira_approved)
    print("not approved jira is",jira_not_approved)
    print("approved jira shape is",jira_approved.shape)
    print("not approved jira shape is",jira_not_approved.shape)
    
    d = pd.DataFrame()
    if d.shape != jira_approved.shape :
        df1=pd.DataFrame()
        approved_jira_df = pd.concat([df1,
        jira_approved[["Assignee"]],
        jira_approved[["Issue Type"]],
        jira_approved[["Story Points"]],
        jira_approved[["Summary"]],
        jira_approved[["Domain"]],
        jira_approved[["Key"]],
        jira_approved[["Complexity"]]
        ],axis=1)
        print("entered the approved if block")
        
        try:
            approved_jira_df.to_excel(consolidate_approved_jiras,sheet_name='data_set',index=False)
        except PermissionError:
            sys.exit("close the opened Excel Sheet "+consolidate_approved_jiras)
        
        #adding the adjustments to thed approved jiras excel sheet
        total_columns=len(approved_jira_df.iloc[0,:])
        total_rows=len(approved_jira_df.iloc[:])
        beautify_excel(consolidate_approved_jiras,total_rows,total_columns)
        file.write("added the adjustment for {} excel file\n".format(consolidate_approved_jiras))
        
        print("total rows in ",consolidate_approved_jiras," are ",len(approved_jira_df.iloc[:])+1,"\n")#1 includes column names
        print("output file name is ",consolidate_approved_jiras)
        
        file.write("total rows in {}  are {}\n".format(consolidate_approved_jiras,len(approved_jira_df.iloc[:])+1))#1 includes column names
        file.write("output file name is {}\n".format(consolidate_approved_jiras))
    
        
        
    d = pd.DataFrame()
    if d.shape != jira_not_approved.shape :
        df1=pd.DataFrame()
        print("entered the not approved if block")
        not_approved_jira_df = pd.concat([df1,
        jira_not_approved[["Assignee"]],
        jira_not_approved[["Issue Type"]],
        jira_not_approved[["Story Points"]],
        jira_not_approved[["Summary"]],
        jira_not_approved[["Domain"]],
        jira_not_approved[["Key"]],
        jira_not_approved[["Complexity"]]
        ],axis=1)
        #print(df1)
        
        try:
            not_approved_jira_df.to_excel(not_approved_consolidated_jiras,sheet_name='data_set',index=False)
        except PermissionError:
            sys.exit("close the opened Excel Sheet "+not_approved_consolidated_jiras)
        #approved_jira_df.to_clipboard(index=False)
        
        
        
        #adding the adjustments to the not approved jiras excel sheet
        total_columns=len(not_approved_jira_df.iloc[0,:])
        total_rows=len(not_approved_jira_df.iloc[:])
        beautify_excel(not_approved_consolidated_jiras,total_rows,total_columns)
        file.write("added the adjustment for {} excel file\n".format(not_approved_consolidated_jiras))
    
    
    
    print("complexes are",complexity_dict)
    s1= pd.Series(complexity_dict)
    try:
        s2 = s1 *[c1_billing,c2_billing,c3_billing,c4_billing,c5_billing]
    except ValueError:
        sys.exit("billing and complex are not matching")
    billings=s2.to_dict()
    file.writelines(["complexes are ", str(complexity_dict)])
    file.writelines(["billing is ",str(billings)])
    print("billings are ",billings)

    print("created the consolidated bill")
    
    
    

if 1:
    
    year=date.today().year
    input_filename=input_file#'WW_40_43_1.xlsx'
    result_file=jira_submission_file#'jira_submission.xlsx'
    try:
        xL = pd.ExcelFile(input_filename)        
    except PermissionError:
        sys.exit("close the opened Excel Sheet"+input_filename)
    print("sheets in excel file are\n",xL.sheet_names)
    list_of_sheets =xL.sheet_names
    print("total no.of sheets are ",len(list_of_sheets)) 
    df = pd.read_excel(input_filename,sheet_name=xL.sheet_names)#accessing sheets by name
    
    df1 = pd.DataFrame()
    
    for i in list_of_sheets:
        print("sheet name is ",i)
        print("columns names in {} sheet are {} ".format(i, df[i].columns.values))
        l=len(df[i].iloc[:])
        for j in range(l):
            if pd.notna(df[i].loc[j,"Assignee"]) and pd.notna(df[i].loc[j,"Intel Leads Approval"]) and df[i].loc[j,"Intel Leads Approval"].strip().lower() == "approved":
                print("approved jira is ",df[i].loc[j,"Key"])
                d = { "PONumber":int(PONumber),
                      "Vendor":Vendor,
                      "Team":Team,
                      "Platform":Platform,
                      "SKU":SKU,
                      "WW":int(workweek),
                      "Year":year,
                      "JiraID":df[i].loc[j,"Key"],
                      "BillableHeader":BillableHeader,
                      "Location":Location,
                      "L1Approver":L1Approver,
                      "L2Approver":L2Approver,
                      "ExcelComment":np.nan,
                      "Custom1":np.nan,
                      "Custom2":np.nan,
                      "ProjectID":np.nan,
                      "BKCID":np.nan
                    }
                ser=pd.Series(d)
                df1=df1.append(ser,ignore_index=True)
    df2=pd.DataFrame()
    d = pd.DataFrame()
    
    if d.shape != df1.shape :
        jira_submission = pd.concat([df2,
                            df1[["PONumber"]],
                            df1[["Vendor"]],
                            df1[["Team"]],
                            df1[["Platform"]],
                            df1[["SKU"]],
                            df1[["WW"]],
                            df1[["Year"]],
                            df1[["JiraID"]],
                            df1[["BillableHeader"]],
                            df1[["Location"]],
                            df1[["L1Approver"]],
                            df1[["L2Approver"]],
                            df1[["ExcelComment"]],
                            df1[["Custom1"]],
                            df1[["Custom2"]],
                            df1[["ProjectID"]],
                            df1[["BKCID"]]
                            ],axis=1)

        try:
            jira_submission.to_excel(result_file,sheet_name='data_set',index=False)
        except PermissionError:
            sys.exit("close the opened Excel Sheet "+result_file)
        
        
        total_columns=len(jira_submission.iloc[0,:])
        total_rows=len(jira_submission.iloc[:])
        #beautify_excel(result_file,total_rows,total_columns)
        
        
        wb = load_workbook(result_file) 
        ws = wb["data_set"]
        #print("total rows are {} and total columns are {}".format(len(jira_submission.iloc[:]),len(jira_submission.iloc[0,:])))

        #applying the colour to the column headings and adjusting the column size to 25 characters
        fill_cell = PatternFill(patternType='solid', fgColor='366092',bgColor='FFFFFF')
        for i in range(1,total_columns+1):
            ws.cell(row=1,column=i).fill =fill_cell
            ws.cell(row=1,column=i).font=Font(color='FFFFFF',bold=True,name='Intel Clear')
            ws.column_dimensions[colToExcel(i)].width = column_width
                    
        #adding the hyperlink to the cell for JIRA
        for i in range(2,total_rows+2):
            ws.cell(row=i, column=8).hyperlink = "https://jira.devtools.intel.com/browse/"+ws.cell(row=i, column=8).value
            ws.cell(row=i, column=8).value = ws.cell(row=i, column=8).value
            ws.cell(row=i, column=8).style = "Hyperlink"
        
        #Align the cells to center
        for i in range(1,total_rows+2):
            for j in range(1,total_columns+1):
                ws.cell(row=i,column=j).alignment = Alignment(horizontal='center', vertical='center')
                
        #save the excel file
        wb.save(result_file)
    
    
        print("total rows in ",result_file," are ",len(jira_submission.iloc[:])+1)#1 includes column names
        print("output file is ",result_file)

file.close()

#path of the chromium driver
#C:\Users\padavenx\Downloads\chromedriver_win32

#update_story_points("ADAD-22774",2)

'''
#renaming the sheet names
xL = pd.ExcelFile('WW_40_43.xlsx')
print("sheets in excel file are\n",xL.sheet_names)
list_of_sheets =xL.sheet_names

with pd.ExcelWriter('WW_40_43_1.xlsx') as writer:
    for i in list_of_sheets:
        df[i].to_excel(writer,sheet_name=i[0:-4],index=False)
'''
