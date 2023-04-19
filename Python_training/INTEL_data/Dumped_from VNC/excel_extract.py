import pandas as pd
import numpy as np
from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill,Alignment
import sys
from datetime import date

print("you have to install the below packages to run")
print("pandas,numpy,openpyxl and xlrd")
print("pip install pandas\npip install numpy\npip install openpyxl\npip install xlrd")
print("python script and input file should be in same folder")
user_request= int(input("which file you wanna generate?\n 1. consolidate  2. JIRA_submission 3. both\n"))

if user_request == 3 or user_request == 1:
    input_filename='WW_40_43_1.xlsx'
    result_filename='consollidated_bill.xlsx'
    xL = pd.ExcelFile(input_filename)
    print("sheets in excel file are\n",xL.sheet_names)
    list_of_sheets =xL.sheet_names
    print("total no.of sheets are ",len(list_of_sheets)) 
    df = pd.read_excel(input_filename,sheet_name=xL.sheet_names)#accessing sheets by name
    
    jira_sumit_df=pd.DataFrame()

    for i in list_of_sheets:
        print("sheet name is ",i)
        l=len(df[i].iloc[:])
        #print("total rows in {0} shett are {1}".format(i,len(df[i].iloc[:,:])))
        for j in range(l):
            if pd.notna(df[i].loc[j,"Assignee"]) and pd.notna(df[i].loc[j,"Intel Leads Approval"]) and df[i].loc[j,"Intel Leads Approval"].strip().lower() == "approved":
            #if pd.notna(df[i].loc[j,"Assignee"]):
                print("approved jira is ",df[i].loc[j,"Key"])
                d = { "Assignee":df[i].loc[j,"Assignee"],
                      "Issue Type":df[i].loc[j,"Issue Type"],
                      "Story Points":df[i].loc[j,"Story Points"],
                      "Summary":df[i].loc[j,"Summary"],
                      "Domain":i,
                      "Key":df[i].loc[j,"Key"],
                      "Complexity":df[i].loc[j,"Complexity"]
                    }
                ser=pd.Series(d)
                jira_sumit_df=jira_sumit_df.append(ser,ignore_index=True)
    

    df1=pd.DataFrame()
    consolidated_df = pd.concat([df1,
    jira_sumit_df[["Assignee"]],
    jira_sumit_df[["Issue Type"]],
    jira_sumit_df[["Story Points"]],
    jira_sumit_df[["Summary"]],
    jira_sumit_df[["Domain"]],
    jira_sumit_df[["Key"]],
    jira_sumit_df[["Complexity"]]    
    ],axis=1)
    consolidated_df.to_excel(result_filename,sheet_name='data_set',index=False)
    
    wb = load_workbook(result_filename) 
    #ws = wb.get_sheet_by_name("data_set")
    ws = wb["data_set"]
    
    #Align the cells to center
    for i in range(1,len(consolidated_df.iloc[:])+2):
        for j in range(1,len(consolidated_df.iloc[0,:])+1):
            ws.cell(row=i,column=j).alignment = Alignment(horizontal='center', vertical='center')
    
    #applying the colour to the column headings
    fill_cell = PatternFill(patternType='solid', fgColor='ffff00')
    for i in range(1,len(consolidated_df.iloc[0,:])+1):
        ws.cell(row=1,column=i).fill =fill_cell
    
    for i in range(2,len(consolidated_df.iloc[:])+2):
        ws.cell(row=i, column=6).hyperlink = "https://jira.devtools.intel.com/browse/"+ws.cell(row=i, column=6).value
        ws.cell(row=i, column=6).value = ws.cell(row=i, column=6).value
        ws.cell(row=i, column=6).style = "Hyperlink"
    
    wb.save(result_filename)
    print("total rows in ",result_filename," are ",len(consolidated_df.iloc[:])+1)#1 includes column names
    print("output file name is ",result_filename)


if user_request == 3 or user_request == 2:
    workweek = int(input("enter the work week"))
    work_week=workweek
    year=date.today().year
    input_filename='WW_40_43_1.xlsx'
    result_file='jira_submission.xlsx'
    xL = pd.ExcelFile(input_filename)
    print("sheets in excel file are\n",xL.sheet_names)
    list_of_sheets =xL.sheet_names
    print("total no.of sheets are ",len(list_of_sheets)) 
    df = pd.read_excel(input_filename,sheet_name=xL.sheet_names)#accessing sheets by name
    
    df1 = pd.DataFrame()
    
    for i in list_of_sheets:
        print("sheet name is ",i)
        l=len(df[i].iloc[:])
        for j in range(l):
            if pd.notna(df[i].loc[j,"Assignee"]) and pd.notna(df[i].loc[j,"Intel Leads Approval"]) and df[i].loc[j,"Intel Leads Approval"].strip().lower() == "approved":
                print("approved jira is ",df[i].loc[j,"Key"])
                d = { "PONumber":int(3002139874),
                      "Vendor":"Cerium",
                      "Team":"E2E - Automation",
                      "Platform":"RAILS",
                      "SKU":"NA",
                      "WW":work_week,
                      "Year":year,
                      "JiraID":df[i].loc[j,"Key"],
                      "BillableHeader":"StoryPointSlab",
                      "Location":"SRR Bangalore",
                      "L1Approver":"Kh, Brinda",
                      "L2Approver":"Jain, Nalina"
                    }
                ser=pd.Series(d)
                df1=df1.append(ser,ignore_index=True)
    df2=pd.DataFrame()
    jira_submission = pd.concat([df2,
                        df1[["PONumber"]],
                        df1[["Vendor"]],
                        df1[["Team"]],
                        df1[["Platform"]],
                        df1[["SKU"]],
                        df1[["WW"]],
                        df1[["Year"]],
                        df1[["JiraID"]],
                        df1[["BillableHeader"]],
                        df1[["Location"]],
                        df1[["L1Approver"]],
                        df1[["L2Approver"]],
                        ],axis=1)
    jira_submission.to_excel(result_file,sheet_name='data_set',index=False)
    wb = load_workbook(result_file) 
    ws = wb.get_sheet_by_name("data_set")
    print("total rows are {} and total columns are {}".format(len(jira_submission.iloc[:]),len(jira_submission.iloc[0,:])))

    #Align the cells to center
    for i in range(1,len(jira_submission.iloc[:])+2):
        for j in range(1,len(jira_submission.iloc[0,:])+1):
            ws.cell(row=i,column=j).alignment = Alignment(horizontal='center', vertical='center')
            
    #applying the colour to the column headings
    fill_cell = PatternFill(patternType='solid', fgColor='ffff00')
    for i in range(1,len(jira_submission.iloc[0,:])+1):
        ws.cell(row=1,column=i).fill =fill_cell
    
    #adding the hyperlink to the cell for JIRA
    for i in range(2,len(jira_submission.iloc[:])+2):
        ws.cell(row=i, column=8).hyperlink = "https://jira.devtools.intel.com/browse/"+ws.cell(row=i, column=8).value
        ws.cell(row=i, column=8).value = ws.cell(row=i, column=8).value
        ws.cell(row=i, column=8).style = "Hyperlink"
    
    #save the excel file
    wb.save(result_file)
    print("total rows in ",result_file," are ",len(jira_submission.iloc[:])+1)#1 includes column names
    print("output file is ",result_file)
