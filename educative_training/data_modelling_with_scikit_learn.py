#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Sep 22 21:17:41 2020

@author: pj
"""
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn import linear_model
from sklearn import tree
from sklearn import metrics
from sklearn.model_selection import train_test_split
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import GridSearchCV
'''
#for creating the sample data
#calories
arr1 = np.random.randint(low=10,high=100,size =(100,1))*100
#weight_in_grams
arr2 = np.random.randint(low=10,high=100,size =(100,1))*10
#price
arr3 = np.random.randint(low=100,high=980,size =(100,1))
pizza_array = np.concatenate([arr1,arr2,arr3],axis=1)
pizza_df = pd.DataFrame(pizza_array,columns=['calories','weight_in_grams','price'])
pizza_df.to_excel('Grouping.xlsx',sheet_name='pizza_data')
print(np.concatenate([arr1,arr2,arr3],axis=1))'''



df = pd.read_excel('Grouping.xls',index_col=0,sheet_name='pizza_data')
pizza_data = df[['calories','weight_in_grams']].values #data set
pizza_price = df['price']#labels

'''#only at start
#creating the data set
data_set = np.random.randint(low=1,high=50,size=(569,4)).astype("float")/10

#creating the labels either 0 or 1
labels = np.random.randint(low=0,high=2,size=(569,1))

df = pd.DataFrame(data_set)
df1 = pd.DataFrame(labels) #from array to dataframe
#stroing data sets and labels in excel file
#Avoiding overwrite
with pd.ExcelWriter('Logistic.xlsx') as writer:
    df.to_excel(writer,sheet_name='data_set')
    df1.to_excel(writer,sheet_name='binary_labels')

#creating the multi labels for the first time
labels = np.random.randint(low=0,high=3,size=(569,1))
df = pd.DataFrame(labels)

#storing labels into excel file
book = load_workbook('Logistic.xlsx')
writer = pd.ExcelWriter('Logistic.xlsx',engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title,ws) for ws in book.worksheets)
df.to_excel(writer,"multi_labels")
writer.save()'''

if 0:#Linear Regression
    if 0:#what is linear Regression
        # finding an optimal model for a dataset is difficult, we instead try to find a good approximating distribution. In many cases, a linear model (a linear combination of the dataset's features) can approximate the data well. The term linear regression refers to using a linear model to represent the relationship between a set of independent variables and a dependent variable.
        
        #Y = ax1+bx2+cx3+d
        pass
    
    if 0:#Basic linear regression
        # The simplest form of linear regression is called least squares regression. This strategy produces a regression model, which is a linear combination of the independent variables, that minimizes the sum of squared residuals between the model's predictions and actual values for the dependent variable.
        
        reg = linear_model.LinearRegression()#calling the regression model
        reg.fit(pizza_data,pizza_price)#preparing the model with our data
        
        #Now the model is ready. we can give new values for testing
        new_pizzas = np.array([[2000,820],[2200,830]])
        print(repr(new_pizzas))
        
        prediceted_prices = reg.predict(new_pizzas)
        print(prediceted_prices)
        
        # Getting the co-efficients and intercept values
        print("co efficients {}\n".format(reg.coef_))
        print("intercept is {}\n".format(reg.intercept_))
        
        # Finally, we can retrieve the "coefficient of determination" (or R2 value) using the score function applied to the dataset and labels. The R2 value tells us how close of a fit the linear model is to the data, or in other words, how good of a fit the model is for the data.
        
        # Using previously defined pizza_data, pizza_prices
        r2 = reg.score(pizza_data, pizza_price)
        print('R2: {}\n'.format(r2))
        
        # The traditional R2 value is a real number between 0 and 1,where lower values denote a poorer model fit to the data. The closer the value is to 1, the better the model's fit on the data.

if 0:#Ridge Regression
    if 1:#
        # While ordinary least squares regression is a good way to fit a linear model onto a dataset, it relies on the fact that the dataset's features are each independent, i.e. uncorrelated. 
        # Because real life data tends to have noise, and will often have some linearly correlated features in the dataset, we combat this by performing regularization.
        # For regularization, the goal is to not only minimize the sum of squared residuals, but to do this with coefficients as small as possible. 
        pass
    
    if 0:#Choosing the best alpha(hyperparameter)
        reg = linear_model.Ridge(alpha=0.1)#manually adding the alphadefault value is 1
        reg.fit(pizza_data,pizza_price)
        print("co-efficients {}\n".format(reg.coef_))
        print("Intercept is {}\n".format(reg.intercept_))
        
        r2 = reg.score(pizza_data,pizza_price)
        print("R2 value is  {}\n".format(r2))
        
    if 1:#Choosing the proper alpha from list of alpha
        alphas = [0.1,0.4,0.6,0.7]
        reg = linear_model.RidgeCV(alphas=alphas)
        reg.fit(pizza_data,pizza_price)
        print("co-efficients {}\n".format(reg.coef_))
        print("Intercept is {}\n".format(reg.intercept_))
        
        r2 = reg.score(pizza_data,pizza_price)
        print("R2 value is  {}\n".format(r2))

if 0:# sparse linear regression via LASSO 
    if 1:# Sparse Regularization
        # Ridge regularization uses L2 norm penalty term but sparse regularization uses L1 norm for the weights penaluty term
        # LASSO regularization tends to prefer linear models with fewer parameter values. This means that it will likely zero-out some of the weight coefficients.
        pass
    
    if 1:#sparse regularization
        data = np.random.uniform(low=1.1,high=5.6,size=(150,4))
        # print(data)
        labels = pd.read_excel('Grouping.xls',sheet_name='Lasso_data')
        # print(labels)
        
        reg = linear_model.Lasso(alpha=0.1)
        reg.fit(data,labels)
        
        print("co-efficients are {}\n".format(reg.coef_))
        print("intercept is {}\n".format(reg.intercept_))
        print("R2 value is {}\n".format(reg.score(data,labels)))
        
if 0:#Bayesian Regression:
    if 1:#Bayesian techniques
        # In Bayesian statistics, the main idea is to make certain assumptions about the probability distributions of a model's parameters before being fitted on data. These initial distribution assumptions are called priors for the model's parameters.
        # In a Bayesian ridge regression model, there are two hyperparameters to optimize: α and λ. The α hyperparameter serves the same exact purpose as it does for regular ridge regression; namely, it acts as a scaling factor for the penalty term.
        # The λ hyperparameter acts as the precision of the model's weights. Basically, the smaller the λ value, the greater the variance between the individual weight values.
        pass
    
    if 1:#Hyperparameter priors
        # Both the α and λ hyperparameters have gamma distribution priors, meaning we assume both values come from a gamma probability distribution.
        # There's no need to know the specifics of a gamma distribution, other than the fact that it's a probability distribution defined by a shape parameter and scale parameter
        pass
    
    if 1:# Tuning the model
        data = np.random.uniform(low=1.1,high=5.6,size=(150,4))
        # print(data)
        labels = pd.read_excel('Grouping.xls',sheet_name='Lasso_data')
        # print(labels)
        
        reg = linear_model.BayesianRidge()
        reg.fit(data,labels)
        
        print('Coefficients: {}\n'.format(repr(reg.coef_)))
        print('Intercept: {}\n'.format(reg.intercept_))
        print('R2: {}\n'.format(reg.score(data, labels)))
        print('Alpha: {}\n'.format(reg.alpha_))
        print('Lambda: {}\n'.format(reg.lambda_))
        
        
if 0:#Logistic regression
    if 0:
        pass
        #         The logistic regression model,
        # despite its name, is actually a linear model for classification. It is called logistic regression because it performs regression on logits (https://en.wikipedia.org/wiki/Logit), 
        # which then allows us to classify the data based on model probability predictions.
        # The default setting for LogisticRegression is binary classification, i.e. classifying data observations that are labeled with either a 0 or 1.
        
    if 0:#binary classification
        '''#only at start
        #creating the data set
        data_set = np.random.randint(low=1,high=50,size=(569,4)).astype("float")/10
        
        #creating the labels either 0 or 1
        labels = np.random.randint(low=0,high=2,size=(569,))
        
        df = pd.DataFrame(data_set)
        df1 = pd.DataFrame(labels) #from array to dataframe
        #stroing data sets and labels in excel file
        #Avoiding overwrite
        with pd.ExcelWriter('Logistic.xlsx') as writer:
            df.to_excel(writer,sheet_name='data_set')
            df1.to_excel(writer,sheet_name='binary_labels')'''
        
        #loading data sets and 
        data_set = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='data_set')
        labels = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='binary_labels')
        
        reg = linear_model.LogisticRegression()
        reg.fit(data_set,labels)
        new_data = np.random.randint(low=1,high=50,size=(2,4)).astype("float")/10
        # print(new_data)
        prediction = reg.predict(new_data)
        print(prediction)
        
    '''For multiclass classification, i.e. when there are more than two labels, we
initialize the LogisticRegression object with the multi_class keyword
argument. The default value is 'ovr' , which signifies a One-Vs-Rest
(https://en.wikipedia.org/wiki/Multiclass_classification#One-vs.-rest)
strategy. In multiclass classification, we want to use the 'multinomial'
strategy.
The code below demonstrates multiclass classification. Note that to use
the 'multinomial' strategy, we need to choose a proper solver (see below
for details on solvers).'''
    
    
    if 0:#multiclass classification
        '''#creating the multi labels for the first time
        labels = np.random.randint(low=0,high=3,size=(569,))
        df = pd.DataFrame(labels)
        
        #storing labels into excel file
        book = load_workbook('Logistic.xlsx')
        writer = pd.ExcelWriter('Logistic.xlsx',engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title,ws) for ws in book.worksheets)
        df.to_excel(writer,"multi_labels")
        writer.save()'''
        
        #loading the data set and labels
        data_set = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='data_set')
        multi_labels = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='multi_labels')
        
        reg0 = linear_model.LogisticRegression(solver='lbfgs',multi_class='multinomial',max_iter=200,penalty='l2')
        
        reg1 = linear_model.LogisticRegression(solver='liblinear',multi_class='ovr',max_iter=300,penalty='l1')
        
        reg0.fit(data_set,multi_labels)
        reg1.fit(data_set,multi_labels)
        
        
        new_data = np.random.randint(low=1,high=50,size=(2,4)).astype("float")/10
        # print(new_data)
        prediction0 = reg0.predict(new_data)
        prediction1 = reg1.predict(new_data)
        
        print(prediction0,'\n',prediction1)
        
    if 1:#cross validated model
        #loading data set and labels
        data_set = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='data_set')
        multi_labels = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='multi_labels')
        reg = linear_model.LogisticRegressionCV(solver='sag',max_iter=400)
        reg.fit(data_set,multi_labels)
        new_data = np.random.randint(low=1,high=50,size=(2,4)).astype("float")/10
        predict = reg.predict(new_data)
        print(predict)
        

if 0:#Decision trees
    pass
    '''It is a binary tree
(https://en.wikipedia.org/wiki/Binary_tree) where each node of the tree
decides on a particular feature of the dataset, and we descend to the
node's left or right child depending on the feature's value.

        If a feature is boolean, we go left or right from the node depending on if the feature value is true or false. If a feature is numeric, we can decide to go left or right based on a decision boundary for the feature (e.g. go left if the feature value is less than 1, otherwise go right).
The leaves of the decision tree determine the class label to predict (in
classification) or the real number value to predict (in regression).'''
    
    if 1:#decision trees
        clf_tree1= tree.DecisionTreeClassifier()
        reg_tree1 = tree.DecisionTreeRegressor()
        clf_tree2 = tree.DecisionTreeClassifier(max_depth=8)
        reg_tree2 = tree.DecisionTreeRegressor(max_depth=8)#no.of layers are 8 i.e max. depth
        
        #loading the data and labels
        data_set = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='data_set')
        labels = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='binary_labels')
        print(data_set.shape)
        print(labels.shape)
        
        #loading the data into model for training
        clf_tree1.fit(data_set,labels)
        new_data = np.random.randint(low=1,high=50,size=(2,4)).astype("float")/10
        
        #verification
        predict = clf_tree1.predict(new_data)
        print(predict)
        # print("co-efficients are {}\n".format(clf_tree1.coef_))
        # print("intercept is {}\n".format(clf_tree1.intercept_))
        #decision tree does not have nay linear equation
        print("R2 value is {}\n".format(clf_tree1.score(data_set,labels)))

if 0:#Splitting the data and labels 
    #loading the data and labels
    data_set = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='data_set')
    labels = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='binary_labels')
    print(data_set.shape)
    print(labels.shape)
    
    #splitting the data set
    split_data = train_test_split(data_set,labels,test_size=0.35)#default test_size=0.25
    train_data = split_data[0]
    test_data = split_data[1]
    train_labels = split_data[2]
    test_labels = split_data[3]
    print(repr(test_labels))

if 0:#cross validation
    '''Each round of the K-Fold algorithm, the model is trained on that round's training set (the combined training folds(k-1 folds)) and then evaluated on the single validation fold(k fold). The evaluation metric depends on the model. For classification models, this is usually classification accuracy on the validation set. For regression models, this can either be the model's mean squared error, mean absolute error, or R 2 value on the validation set.'''
    
    if 0:#classification with K-Fold CV with 3 folds
        clf = linear_model.LogisticRegression()
        
        #loading the data and labels
        data_set = pd.read_excel('0Logistic.xlsx',index_col=0,sheet_name='data_set')
        print("data set is {}\n".format(type(data_set)))
        
        df = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='binary_labels')
        labels = pd.get_dummies(df).values.ravel()#ravel using for removal of warnings
        print("labels is {}\n".format(type(labels)))
        
        cv_score = cross_val_score(clf,data_set,labels,cv=3)# 3 folds
        print("{}\n".format(repr(cv_score)))
        
    if 1:#Regression with K-Fold CV with 4 folds
        clf = linear_model.LinearRegression()
        
        #loading data labels
        df = pd.read_excel('Grouping.xlsx',index_col=0,sheet_name='pizza_data')
        data_set = df[['calories','weight_in_grams']].values #data set
        labels = df['price']#labels
        
        cv_score = cross_val_score(clf,data_set,labels,cv=4)# 4 folds
        print("{}\n".format(repr(cv_score)))
        

if 0:#Applying the CV to the decision trees to get proper max_depth value
    def cv_decicion_tree(is_clf,data,labels,folds,maximum_depth):
        if is_clf:
            d_tree = tree.DecisionTreeClassifier(max_depth=maximum_depth)
        else :
            d_tree = tree.DecisionTreeRegressor(max_depth=maximum_depth)
        
        cv_scores = cross_val_score(d_tree,data,labels,cv=folds)
        return cv_scores
    
    if 0:#for classification
        #loading the data and labels    
        data_set = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='data_set')
        labels = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='binary_labels')
        
        #calling the method to apply cv for classification
        for depth in range(3,8):
            scores = cv_decicion_tree(True, data_set, labels, 5, depth)
            mean = scores.mean() #mean across the folds 
            std_2 = 2 *scores.std() #std deviations
            print("depth is {},mean is {}, std dev is +/-{:.2f}".format(depth,mean,std_2))
    
    if 1:#For regression
        #loading data and labels
        df = pd.read_excel('Grouping.xlsx',index_col=0,sheet_name='pizza_data')
        data = df[['calories','weight_in_grams']]
        data_set = pd.get_dummies(data).values
        
        label=df['price']
        labels = pd.get_dummies(label).values
        # print(type(labels))
        
        #calling the method to apply cv for classification
        for depth in range(3,8):
            scores = cv_decicion_tree(False, data_set, labels, 5, depth)
            mean = scores.mean() #mean across the folds 
            std_2 = 2 *scores.std() #std deviations
            print("depth is {},mean is {}, std dev is +/-{:.2f}".format(depth,mean,std_2))

if 0:#Evaluating models
    '''For classification models, we use the classification accuracy on the test set as the evaluation metric. For regression models, we normally use either the R 2 value, mean squared error, or mean absolute error on the test set. The most commonly used regression metric is mean absolute error, since it represents the natural definition of error. We use mean squared error when we want to penalize really bad predictions, since the error is squared. We use the R 2 value when we want to evaluate the fit of the regression model on the data'''
    
    if 1:#Evaluating for classification model
        #loading the data
        data_set = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='data_set')
        data_set = pd.get_dummies(data_set).values
        
        labels = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='binary_labels')
        labels = pd.get_dummies(labels).values
        
        #splitting into training and testing
        split_data = train_test_split(data_set, labels)
        train_data = split_data[0]
        test_data = split_data[1]
        train_labels = split_data[2]
        test_labels = split_data[3]
        
        #training the model
        clf = tree.DecisionTreeClassifier()
        clf.fit(train_data, train_labels)
        predictions = clf.predict(test_data)
        
        #Evaluation metrics
        accuracy = metrics.accuracy_score(test_labels, predictions)
        print("Accuracy is {}\n".format(accuracy))
        
    if 0:#Evaluating for regression model
        #loading the data
        df = pd.read_excel('Regression.xlsx',index_col=0,sheet_name='pizza_data')
        data_set = df[['calories','weight_in_grams']]
        data_set = pd.get_dummies(data_set).values
        
        labels = df['price']
        labels = pd.get_dummies(labels).values
        
        #splitting into training and testing
        split_data = train_test_split(data_set, labels)
        train_data = split_data[0]
        test_data = split_data[1]
        train_labels = split_data[2]
        test_labels = split_data[3]
        
        #training the model
        reg = tree.DecisionTreeRegressor()
        reg.fit(train_data,train_labels)
        predictions = reg.predict(test_data)
        
        
        #Evaluation metrics
        r2 = metrics.r2_score(test_labels,predictions)#how muach fitted
        print("R2 is {}\n".format(r2))
        
        mse = metrics.mean_squared_error(test_labels,predictions)
        print("MSE is {}\n".format(mse))#to penalise the bad predictions
        
        mae = metrics.mean_absolute_error(test_labels,predictions)
        print("MAE is {}\n".format(mae))

if 1:#Exhaustive Tuning
    reg = linear_model.BayesianRidge()
    #loading the data 
    data_set = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='data_set')
    # data_set = pd.get_dummies(data_set).values.ravel()
    labels = pd.read_excel('Logistic.xlsx',index_col=0,sheet_name='binary_labels')
    labels = pd.get_dummies(labels).values.ravel()
    
    #selecting the hyperparams
    params = {'alpha_1':[0.1,0.2,0.3],
              'alpha_2':[0.1,0.2,0.3]}
    #splitting into training and testing
    split_data = train_test_split(data_set, labels)
    train_data = split_data[0]
    test_data = split_data[1]
    train_labels = split_data[2]
    test_labels = split_data[3]
    
    #training the model
    reg_cv = GridSearchCV(reg,params,cv=5,iid=False)#5 folds
    reg_cv.fit(train_data,train_labels)
    print(reg_cv.best_params_)












