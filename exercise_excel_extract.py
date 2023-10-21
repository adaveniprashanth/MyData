import configparser
import pandas as pd
import numpy as np
import xlsxwriter
import xlrd
from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill,Alignment
from openpyxl.utils import get_column_letter
import sys,re
from datetime import date
import time
from selenium import webdriver
from selenium.webdriver.common.by import By

print("you have to install the below packages to run")
print("pandas,numpy,openpyxl and xlrd")
print("pip install pandas\npip install numpy\npip install openpyxl\npip install xlrd")
print("python script and input file should be in same folder")

#importing from the config gile
config=configparser.ConfigParser()
config.read("configurations.ini") #reading the config file
input_file= config["FILESettings"]["input_filename"]
consolidated_file=config["FILESettings"]["consolidated_filename"]
jira_submission_file = config["FILESettings"]["jira_submission_filename"]
workweek=config["FILESettings"]["work_week"]
PONumber=config["FILESettings"]["PONumber"]
Vendor=config["FILESettings"]["Vendor"]
Team = config["FILESettings"]["Team"]
Platform = config["FILESettings"]["Platform"]
SKU = config["FILESettings"]["SKU"]
BillableHeader =config["FILESettings"]["BillableHeader"]
Location = config["FILESettings"]["Location"]
L1Approver =config["FILESettings"]["L1Approver"]
L2Approver = config["FILESettings"]["L2Approver"]
column_width= config["FILESettings"]["column_width"]

Assignee=0
Intel_Leads_Approval=12
Key=9
Issue_Type=1
Story_Points=6
Summary=7
Complexity=11
Remarks=13

def update_story_points(jira_ID,expected_point):
    driver = webdriver.Chrome("C:\chromedriver\chromedriver.exe")
    url="https://jira.devtools.intel.com/browse/"
    expected_story_points = expected_point  # find from excel and convert to int
    jira_id = jira_ID
    driver.get("{}{}".format(url,jira_id))
    driver.maximize_window()
    time.sleep(10)
    story_points = int(driver.find_element(By.XPATH, '//div[@class="value type-float editable-field inactive"]').text)
    if story_points != expected_story_points:
        print("update story points")
        driver.find_element(By.XPATH, '//span [@class="icon aui-icon aui-icon-small aui-iconfont-edit"]').click()
        time.sleep(10)
        driver.find_element(By.ID, "customfield_11204").clear()
        driver.find_element(By.ID, "customfield_11204").send_keys(expected_story_points)
        driver.find_element(By.ID, "edit-issue-submit").click()
    else:
        print("story already points set to {}".format(expected_story_points))

#convert column numbers to column names
def colToExcel(col): # col is 1 based
    excelCol = str()
    div = col
    while div:
        (div, mod) = divmod(div-1, 26) # will return (x, 0 .. 25)
        excelCol = chr(mod + 65) + excelCol
    return excelCol

def beautify_excel(file_name):
    #convert column numbers to column names
    def colToExcel(col): # col is 1 based
        excelCol = str()
        div = col
        while div:
            (div, mod) = divmod(div-1, 26) # will return (x, 0 .. 25)
            excelCol = chr(mod + 65) + excelCol

        return excelCol
    
    wb = load_workbook(file_name) 
    ws = wb["data_set"]
    #ws.column_dimensions['A'].width = 25
    
    #applying the colour to the column headings and adjusting the columns width 
    fill_cell = PatternFill(patternType='solid', fgColor='ffff00')
    for i in range(1,len(consolidated_df.iloc[0,:])+1):
        ws.cell(row=1,column=i).fill =fill_cell
        ws.column_dimensions[colToExcel(i)].width = column_width
    
    #adding hyperlink
    for i in range(2,len(consolidated_df.iloc[:])+2):
        ws.cell(row=i, column=6).hyperlink = "https://jira.devtools.intel.com/browse/"+ws.cell(row=i, column=6).value
        ws.cell(row=i, column=6).value = ws.cell(row=i, column=6).value
        ws.cell(row=i, column=6).style = "Hyperlink"
    
    #Align the cells to center
    for i in range(1,len(consolidated_df.iloc[:])+2):
        for j in range(1,len(consolidated_df.iloc[0,:])+1):
            ws.cell(row=i,column=j).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

    print("rows are",ws.rows)
    wb.save(file_name)


if 1:
    input_filename=input_file#'WW_40_43_1.xlsx'
    result_filename=consolidated_file#'consollidated_bill.xlsx'
    xL = pd.ExcelFile(input_filename)
    print("sheets in excel file are\n",xL.sheet_names)
    list_of_sheets =xL.sheet_names
    print("total no.of sheets are ",len(list_of_sheets)) 
    df = pd.read_excel(input_filename,sheet_name=xL.sheet_names)#accessing sheets by name
    
    jira_sumit_df=pd.DataFrame()

    for i in list_of_sheets:
        print("sheet name is ",i)
        l=len(df[i].iloc[:])
        #print("total rows in {0} shett are {1}".format(i,len(df[i].iloc[:,:])))
        for j in range(l):
            if pd.notna(df[i].iloc[j,Assignee]) and pd.notna(df[i].iloc[j,Intel_Leads_Approval]) and df[i].iloc[j,Intel_Leads_Approval].strip().lower() == "approved":
                #if pd.notna(df[i].iloc[j,Remarks]):
                #    update_story_points(df[i].iloc[j,Key],int(re.findall("[0-9]",df[i].iloc[j,Remarks])[0]))
                print("approved jira is ",df[i].iloc[j,Key])
                #print("approved jira is ",df[i].iloc[j,Key])
                d = { "Assignee":df[i].iloc[j,Assignee],
                      "Issue Type":df[i].iloc[j,Issue_Type],
                      "Story Points":df[i].iloc[j,Story_Points],
                      "Summary":df[i].iloc[j,Summary],
                      "Domain":i,
                      "Key":df[i].iloc[j,Key],
                      "Complexity":df[i].iloc[j,Complexity]
                    }
                ser=pd.Series(d)
                jira_sumit_df=jira_sumit_df.append(ser,ignore_index=True)
    

    df1=pd.DataFrame()
    consolidated_df = pd.concat([df1,
                                    jira_sumit_df[["Assignee"]],
                                    jira_sumit_df[["Issue Type"]],
                                    jira_sumit_df[["Story Points"]],
                                    jira_sumit_df[["Summary"]],
                                    jira_sumit_df[["Domain"]],
                                    jira_sumit_df[["Key"]],
                                    jira_sumit_df[["Complexity"]]    
                                ],axis=1)
    consolidated_df.to_excel(result_filename,sheet_name='data_set',index=False)
    consolidated_df.to_clipboard(index=False)
    beautify_excel(result_filename)

    print("total rows in ",result_filename," are ",len(consolidated_df.iloc[:])+1)#1 includes column names
    print("output file name is ",result_filename)
    

if 0:
    
    year=date.today().year
    input_filename=input_file#'WW_40_43_1.xlsx'
    result_file=jira_submission_file#'jira_submission.xlsx'
    xL = pd.ExcelFile(input_filename)
    print("sheets in excel file are\n",xL.sheet_names)
    list_of_sheets =xL.sheet_names
    print("total no.of sheets are ",len(list_of_sheets)) 
    df = pd.read_excel(input_filename,sheet_name=xL.sheet_names)#accessing sheets by name
    
    df1 = pd.DataFrame()
    
    for i in list_of_sheets:
        print("sheet name is ",i)
        l=len(df[i].iloc[:])
        for j in range(l):
            if pd.notna(df[i].loc[j,"Assignee"]) and pd.notna(df[i].loc[j,"Intel Leads Approval"]) and df[i].loc[j,"Intel Leads Approval"].strip().lower() == "approved":
                print("approved jira is ",df[i].loc[j,"Key"])
                d = { "PONumber":int(PONumber),
                      "Vendor":Vendor,
                      "Team":Team,
                      "Platform":Platform,
                      "SKU":SKU,
                      "WW":int(workweek),
                      "Year":year,
                      "JiraID":df[i].loc[j,"Key"],
                      "BillableHeader":BillableHeader,
                      "Location":Location,
                      "L1Approver":L1Approver,
                      "L2Approver":L2Approver
                    }
                ser=pd.Series(d)
                df1=df1.append(ser,ignore_index=True)
    df2=pd.DataFrame()
    jira_submission = pd.concat([df2,
                        df1[["PONumber"]],
                        df1[["Vendor"]],
                        df1[["Team"]],
                        df1[["Platform"]],
                        df1[["SKU"]],
                        df1[["WW"]],
                        df1[["Year"]],
                        df1[["JiraID"]],
                        df1[["BillableHeader"]],
                        df1[["Location"]],
                        df1[["L1Approver"]],
                        df1[["L2Approver"]],
                        ],axis=1)
    jira_submission.to_excel(result_file,sheet_name='data_set',index=False)
    
    wb = load_workbook(result_file) 
    ws = wb["data_set"]
    print("total rows are {} and total columns are {}".format(len(jira_submission.iloc[:]),len(jira_submission.iloc[0,:])))

    

    #applying the colour to the column headings and adjusting the column size to 25 characters
    fill_cell = PatternFill(patternType='solid', fgColor='ffff00')
    for i in range(1,len(jira_submission.iloc[0,:])+1):
        ws.cell(row=1,column=i).fill =fill_cell
        ws.column_dimensions[colToExcel(i)].width = 25
        
        
    #adding the hyperlink to the cell for JIRA
    for i in range(2,len(jira_submission.iloc[:])+2):
        ws.cell(row=i, column=8).hyperlink = "https://jira.devtools.intel.com/browse/"+ws.cell(row=i, column=8).value
        ws.cell(row=i, column=8).value = ws.cell(row=i, column=8).value
        ws.cell(row=i, column=8).style = "Hyperlink"

    
    #Align the cells to center
    for i in range(1,len(jira_submission.iloc[:])+2):
        for j in range(1,len(jira_submission.iloc[0,:])+1):
            ws.cell(row=i,column=j).alignment = Alignment(horizontal='center', vertical='center')
            
    
    #save the excel file
    wb.save(result_file)
    print("total rows in ",result_file," are ",len(jira_submission.iloc[:])+1)#1 includes column names
    print("output file is ",result_file)


#path of the chromium driver
#C:\Users\padavenx\Downloads\chromedriver_win32



#update_story_points("ADAD-22774",2)